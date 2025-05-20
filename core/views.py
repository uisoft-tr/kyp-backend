from rest_framework import viewsets, status
from .models import (
    Makine,
    MakineTakip,
    Planlamalar,
    Yatirim,
    KanunMaddeleri,
    KanunDosyalari,
    Ilce,
    Il,
    IsTanimi,
    ModelAdi,
    Marka,IsinAdi
)
from .serializers import (
    MakineSerializer,
    MakineTakipSerializer,
    PlanlamalarSerializer,
    YatirimSerializer,
    KanunSerializer,
    IlSerializer,
    IlceSerializer,
    IsTanimiSerializer,
    CustomTokenObtainPairSerializer,
    ModelAdiSerializer,
    MarkaSerializer,IsinAdiSerializer
)
from django.contrib.gis.geos import Point
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .filters import MakineTakipFilter, IlceFilter, MakineFilter, PlanlamaFilter,YatirimFilter,KanunFilter
from rest_framework.decorators import api_view, action
from django.db.models import (
    Sum,
    F,
    ExpressionWrapper,
    DurationField,
    OuterRef,
    Subquery,
    Case,
    When,
    Value,
    CharField,
)
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.gis.db.models.functions import Transform
import numpy as np
import json
from sklearn.cluster import DBSCAN
from django.contrib.gis.geos import Point, Polygon
from django.contrib.gis.gdal import SpatialReference, CoordTransform
from openpyxl import Workbook
import datetime
from django.http import HttpResponse
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook
from rest_framework.parsers import MultiPartParser,FormParser,JSONParser
import openpyxl
import traceback
from openpyxl.styles import Border, Side,numbers

class IsTanimiViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = IsTanimi.objects.all().order_by('id')  
    serializer_class = IsTanimiSerializer
    pagination_class = None 

class IsinAdiViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = IsinAdi.objects.all().order_by('id')  
    serializer_class = IsinAdiSerializer
    pagination_class = None 
    
class IlViewSet(viewsets.ModelViewSet):
    queryset = Il.objects.all().order_by('id')  
    serializer_class = IlSerializer
    pagination_class = None 


    @action(detail=False, methods=["get"])
    def ozet(self, request):
        iller = self.get_queryset()
        ozet = [{"id": il.id, "ad": il.name} for il in iller]
        return Response(ozet)

class IlceViewSet(viewsets.ModelViewSet):
    queryset = Ilce.objects.all().order_by('id')  
    serializer_class = IlceSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = IlceFilter
    pagination_class = None 


    def get_queryset(self):
        il_id = self.request.query_params.get("il_id")
        if il_id:
            return Ilce.objects.filter(il_id=il_id)
        return Ilce.objects.all()

class MarkaViewSet(viewsets.ModelViewSet):
    queryset = Marka.objects.all().order_by('id')  
    serializer_class = MarkaSerializer
    pagination_class = None 

class ModelAdiViewSet(viewsets.ModelViewSet):
    queryset = ModelAdi.objects.all().order_by('id')  
    serializer_class = ModelAdiSerializer
    pagination_class = None 

class MakineViewSet(viewsets.ModelViewSet):
    serializer_class = MakineSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = MakineFilter

    def get_queryset(self):
        # En son takip kaydının iş durumunu al
        latest_takip = MakineTakip.objects.filter(makina=OuterRef("pk")).order_by("-id")

        return (
            Makine.objects.annotate(
                son_is_durumu=Subquery(latest_takip.values("isin_durumu")[:1])
            )
            .annotate(
                # Sıralama için önce devam sonra bitti
                oncelik=Case(
                    When(son_is_durumu="devam", then=Value(0)),
                    When(son_is_durumu="bitti", then=Value(1)),
                    default=Value(2),
                    output_field=CharField(),
                )
            )
            .order_by("oncelik", "no")
            .prefetch_related("takipler")
        )

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        filter = self.filterset_class(request.query_params, queryset=queryset)
        no_pagination = request.query_params.get("no_pagination", "false") == "true"
        if no_pagination:
            serializer = self.get_serializer(filter.qs, many=True)
            return Response(serializer.data)
        page = self.paginate_queryset(filter.qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(filter.qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        filter = self.filterset_class(request.query_params, queryset=queryset)
        queryset = filter.qs

        wb = Workbook()
        ws = wb.active
        ws.title = "Makineler"

        # Başlıklar
        ws.append([
            "Makine No",
            "Model Adı",
            "Tür",
            "Cins",
            "Marka",
            "Ait Olduğu Yer",
            "Çalışma Durumu"
        ])

        for obj in queryset:
            ws.append([
                obj.no,
                obj.model_adi.ad if obj.model_adi else "",
                obj.tur or "",
                obj.cins or "",
                obj.marka.ad if obj.marka else "",
                obj.ait_oldugu_yer,
                "Çalışmıyor" if obj.calisma_durumu == "bitti" else "Devam ediyor" if obj.calisma_durumu == "devam" else "",
            ])

        # Hücre kenarlıkları (renksiz çizgili tablo)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        # Excel dosyasını response olarak döndür
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"makineler_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class MakineTakipViewSet(viewsets.ModelViewSet):
    queryset = MakineTakip.objects.all().order_by('id')  
    serializer_class = MakineTakipSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = MakineTakipFilter

    def create(self, request, *args, **kwargs):
        data = request.data.copy()

        # 1. Konum verisini işle
        konum_data = data.get("konum")
        if isinstance(konum_data, dict) and "coordinates" in konum_data:
            try:
                lng, lat = konum_data["coordinates"]
                data["konum"] = Point(float(lng), float(lat))
            except Exception as e:
                return Response(
                    {"error": "Geçersiz konum verisi."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 2. is_tanimi alanını işle
        is_tanimi_value = data.get("is_tanimi")
        if isinstance(is_tanimi_value, str):
            is_tanimi_obj, _ = IsTanimi.objects.get_or_create(tanim=is_tanimi_value)
            data["is_tanimi"] = is_tanimi_obj.id  # ForeignKey olarak ID gönderilmeli

        # 3. Diğer işlemler
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


    def update(self, request, *args, **kwargs):
        data = request.data.copy()

        # 1. Konum verisini işle
        konum_data = data.get("konum")
        if isinstance(konum_data, dict) and "coordinates" in konum_data:
            try:
                lng, lat = konum_data["coordinates"]
                data["konum"] = Point(float(lng), float(lat))
            except Exception as e:
                return Response(
                    {"error": "Geçersiz konum verisi."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 2. is_tanimi alanını işle
        is_tanimi_value = data.get("is_tanimi")
        if isinstance(is_tanimi_value, str):
            is_tanimi_obj, _ = IsTanimi.objects.get_or_create(tanim=is_tanimi_value)
            data["is_tanimi"] = is_tanimi_obj.id  # ForeignKey olarak ID gönderilmeli

        # 3. Güncelleme işlemi
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def makina_calistigi_yer(self, request):
        makina_id = request.query_params.get("makina_id")

        if not makina_id:
            return Response({"error": "Makine ID parametresi gereklidir."}, status=400)

        queryset = MakineTakip.objects.filter(makina_id=makina_id)

        filterset = MakineTakipFilter(request.GET, queryset=queryset)
        if not filterset.is_valid():
            return Response(filterset.errors, status=400)

        filtered_qs = filterset.qs

        result = []
        total_calisma_suresi = 0

        for takip in filtered_qs:
            if takip.ise_bitis:
                calisma_suresi = (takip.ise_bitis - takip.ise_baslama).days
            else:
                calisma_suresi = 0

            total_calisma_suresi += calisma_suresi

            result.append(
                {
                    "id": takip.id, 
                    "makine_no": takip.makina.no, 
                    "calistigi_yer": takip.calistigi_yer,
                    "ise_baslama": takip.ise_baslama,
                    "ise_bitis": takip.ise_bitis,
                    "is_tanimi_ad": str(takip.is_tanimi),
                    "calisma_suresi": calisma_suresi,
                }
            )

        export_format = request.query_params.get("export")
        if export_format == "excel":
            print("Excel export başlıyor...")  # Debug 1
            wb = Workbook()
            ws = wb.active
            ws.title = "MakineDetay"

            headers = ["Makine No", "Calıştığı Yer", "İşe Başlama", "İşe Bitiş", "İş Tanımı", "Çalışma Süresi (Gün)"]
            ws.append(headers)
            print("Başlıklar eklendi.")  # Debug 2

        # Veri satırları
            for item in result:
                ws.append([
                    item["makine_no"],
                    item["calistigi_yer"],
                    item["ise_baslama"].strftime("%Y-%m-%d") if item["ise_baslama"] else "",
                    item["ise_bitis"].strftime("%Y-%m-%d") if item["ise_bitis"] else "",
                    item["is_tanimi_ad"],
                    item["calisma_suresi"],
                ])
                
            print("Tüm veriler eklendi.")  # Debug 4
            last_row = ws.max_row
            last_column = ws.max_column
            table_ref = f"A1:{chr(64 + last_column)}{last_row}" 
            
            print(f"Tablo aralığı: {table_ref}")  # Debug 5

        
            tab = Table(displayName="ÇalışmalarTablosu", ref=table_ref)
            ws.add_table(tab)
            
            print("Tablo eklendi.")  # Debug 6

        
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
            
            print("Borderlar eklendi.")  # Debug 7


            # Excel dosyasını HttpResponse ile döndür
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"makina_calistigi_yer_{makina_id}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            print(f"Dosya adı: {filename}")  # Debug 8

            wb.save(response)
            print("Excel dosyası kaydedildi ve gönderiliyor.")  # Debug 9

            return response

        # Sayfalama uygula
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(result, request, view=self)
        response = paginator.get_paginated_response(page)

     # Toplam çalışma süresi ekle
        response.data["toplam_calisma_suresi"] = total_calisma_suresi

        return response

    @action(detail=False, methods=["get"], url_path="calisan-makineler")
    def calisan_makineler(self, request):
        try:
            queryset = self.get_queryset().filter(isin_durumu="devam")

            paginator = self.pagination_class()
            page = paginator.paginate_queryset(queryset, request, view=self)

            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return paginator.get_paginated_response(serializer.data)

            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="calisan_makineler_geometry")
    def calisan_makineler_geometry(self, request):
        try:
            bbox = request.GET.get("bbox", None)
            zoom = int(float(request.GET.get("zoom", 10)))
            queryset = (
                self.get_queryset().filter(isin_durumu="devam").select_related("makina")
            )

            if bbox:
                bbox_coords = list(map(float, bbox.split(",")))
                minx, miny, maxx, maxy = bbox_coords

                # EPSG:3857 → EPSG:4326 dönüşüm
                source_srs = SpatialReference(3857)
                target_srs = SpatialReference(4326)
                transform = CoordTransform(source_srs, target_srs)

                min_point = Point(minx, miny, srid=3857)
                max_point = Point(maxx, maxy, srid=3857)
                min_point.transform(transform)
                max_point.transform(transform)

                minx, miny = min_point.x, min_point.y
                maxx, maxy = max_point.x, max_point.y

                bbox_polygon = Polygon.from_bbox((minx, miny, maxx, maxy))
                bbox_polygon.srid = 4326

                queryset = queryset.filter(konum__intersects=bbox_polygon)

            queryset = queryset.annotate(geom=Transform("konum", 4326))

            if zoom < 10:
                coordinates = []
                for obj in queryset:
                    centroid = obj.konum.centroid
                    coordinates.append([centroid.x, centroid.y])

                coordinates = np.array(coordinates)
                if len(coordinates) == 0:
                    return Response({"type": "FeatureCollection", "features": []})

                db = DBSCAN(eps=0.7, min_samples=1).fit(coordinates)
                labels = db.labels_
                unique_labels = set(labels)

                response_data = {"type": "FeatureCollection", "features": []}
                for label in unique_labels:
                    cluster_indices = [i for i, l in enumerate(labels) if l == label]
                    cluster_points = [coordinates[i] for i in cluster_indices]
                    cluster_centroid = np.mean(cluster_points, axis=0)
                    cluster_size = len(cluster_points)

                    makina_nolar = set()
                    is_bilgileri = []

                    for i in cluster_indices:
                        takip_obj = queryset[i]
                        makina_obj = takip_obj.makina
                        if makina_obj:
                            makina_nolar.add(makina_obj.no)

                        is_bilgileri.append(
                            {
                                "ise_baslama": takip_obj.ise_baslama,
                                "calistigi_yer": takip_obj.calistigi_yer,
                                "is_tanimi": str(takip_obj.is_tanimi),
                            }
                        )

                    feature = {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [cluster_centroid[0], cluster_centroid[1]],
                        },
                        "properties": {
                            "cluster_label": int(label),
                            "cluster_size": cluster_size,
                            "makina_no": (
                                ", ".join(makina_nolar) if makina_nolar else None
                            ),
                            "is_bilgileri": is_bilgileri,
                        },
                    }
                    response_data["features"].append(feature)

                return Response(response_data)

            else:
                # Zoom 10 ve üzeri için her nokta ayrı döner, iş bilgileri ekleniyor
                features = []
                for obj in queryset:
                    is_bilgileri = (
                        [
                            {
                                "ise_baslama": obj.ise_baslama,
                                "calistigi_yer": obj.calistigi_yer,
                                "is_tanimi": str(obj.is_tanimi),
                            }
                        ]
                        if obj.ise_baslama or obj.calistigi_yer or obj.is_tanimi
                        else []
                    )

                    features.append(
                        {
                            "type": "Feature",
                            "geometry": json.loads(obj.konum.geojson),
                            "properties": {
                                "makina_no": obj.makina.no if obj.makina else None,
                                "is_bilgileri": is_bilgileri,
                            },
                        }
                    )

                response_data = {"type": "FeatureCollection", "features": features}

                return Response(response_data)

        except Exception as e:
            return Response({"error": str(e)}, status=400)

class PlanlamalarViewSet(viewsets.ModelViewSet):
    queryset = Planlamalar.objects.all().order_by("id")
    serializer_class = PlanlamalarSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = PlanlamaFilter
    parser_classes = [MultiPartParser, FormParser,JSONParser] # Dosya yüklemek için


    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.get_queryset()
        queryset = self.filter_queryset(queryset)  

        wb = Workbook()
        ws = wb.active
        ws.title = "Planlamalar"

        ws.append([
            "Sıra No", "Bölge No", "İl", "İlçe", "Taşkın İşi Adı",
            "İlk İnceleme", "Ön İnceleme", "Kamulaştırma",
            "Yaklaşık Maliyet", "Korunan Yerleşim", "Açıklama"
        ])

        for obj in queryset:
            ws.append([
                obj.sira_no,
                obj.bolge_no,
                obj.il.name if obj.il else "",
                obj.ilce.name if obj.ilce else "",
                obj.taskin_isi_adi,
                obj.ilk_inceleme_raporu,
                obj.on_inceleme_raporu,
                obj.kamulastirma_problemi,
                float(obj.yaklasik_insaat_maliyeti),
                obj.korunan_yerlesim_yeri,
                obj.aciklama,
            ])
            
        # Burada tablo oluşturuyoruz:
        last_row = ws.max_row
        last_column = ws.max_column
        table_ref = f"A1:{chr(64 + last_column)}{last_row}"  # Örn: "A1:K20"
        
        tab = Table(displayName="PlanlamalarTablosu", ref=table_ref)
        ws.add_table(tab)
        
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        file_name = f"planlamalar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import_excel')
    def import_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "Excel dosyası yüklenmedi."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # İlk satır başlık, satırları döngü ile işle
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            data_rows = rows[1:]


            for row in data_rows:
                # İl ve İlçe objelerini isim ile çekmek lazım
                il_name = row[2]
                ilce_name = row[3]

                il_obj = None
                if il_name:
                    il_obj = Il.objects.filter(name=il_name).first()

                ilce_obj = None
                if ilce_name:
                    ilce_obj = Ilce.objects.filter(name=ilce_name).first()

                # Satırdaki verilerle Planlamalar objesi oluştur
                Planlamalar.objects.create(
                    sira_no=row[0] or 0,
                    bolge_no=row[1] or 0,
                    il=il_obj,
                    ilce=ilce_obj,
                    taskin_isi_adi=row[4] or "",
                    ilk_inceleme_raporu=row[5] or "yok",
                    on_inceleme_raporu=row[6] or "yok",
                    kamulastirma_problemi=row[7] or "yok",
                    yaklasik_insaat_maliyeti=row[8] or 0,
                    korunan_yerlesim_yeri=row[9] or "",
                    aciklama=row[10] or "",
                )

            return Response({"success": "Excel dosyasından planlamalar başarıyla eklendi."})
        except Exception as e:
            print(traceback.format_exc())  # konsola detaylı trace verir
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class YatirimViewSet(viewsets.ModelViewSet):
    queryset = Yatirim.objects.all().order_by("id")
    serializer_class = YatirimSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = YatirimFilter
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def create(self, request, *args, **kwargs):
        data = request.data.copy()

        isin_adi_value = data.get("isin_adi")
        if isinstance(isin_adi_value, str):
            isin_adi_obj, _ = IsinAdi.objects.get_or_create(ad=isin_adi_value)
            data["isin_adi"] = isin_adi_obj.id  

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


    def update(self, request, *args, **kwargs):
        data = request.data.copy()

        isin_adi_value = data.get("isin_adi")
        if isinstance(isin_adi_value, str):
            isin_adi_obj, _ = IsinAdi.objects.get_or_create(ad=isin_adi_value)
            data["isin_adi"] = isin_adi_obj.id  

        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.get_queryset()
        queryset = self.filter_queryset(queryset)

        wb = Workbook()
        ws = wb.active
        ws.title = "Yatırımlar"

        ws.append([
            "İşin Adı", "Toplam Keşif ve İhale Bedeli", "Yıl Sonuna Kadar Harcama", "Yıl Keşif Bedeli",
            "Yıl Nakdi", "Revize Ödenek", "BBB ve Sonrası Keşif Bedeli",
            "Yapılış Şekli", "Başlama Yılı", "Bitiş Yılı",
            "Talep", "Tenkis"
        ])
        for obj in queryset:
            yapilis_sekli_display = (
            "Kamulaştırma" if obj.yapilis_sekli == "kamulastırma" else
            "Kamulaştırma Dışı" if obj.yapilis_sekli == "kamulastırma_disi" else
            obj.yapilis_sekli
            )

            ws.append([
                obj.isin_adi.ad,
                float(obj.toplam_kesif_ihale_bedeli),
                float(obj.yil_sonuna_kadar_harcama),
                float(obj.yil_kesif_bedeli),
                float(obj.yil_nakti),
                float(obj.revize_odenk),
                float(obj.bbb_ve_sonrasi_kesif_bedeli),
                yapilis_sekli_display,
                obj.baslama_tarihi,
                obj.bitis_tarihi,
                float(obj.talep),
                float(obj.tenkis),
            ])

        last_row = ws.max_row
        last_column = ws.max_column
        table_ref = f"A1:{chr(64 + last_column)}{last_row}"

        tab = Table(displayName="YatirimlarTablosu", ref=table_ref)
        ws.add_table(tab)
        
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for row in ws.iter_rows(min_row=2, max_row=last_row):
            for col in range(2, 8):
                cell = row[col - 1]
                cell.number_format = '#,##0.00'
        for col in range(11, 13):
            cell = row[col - 1]
            cell.number_format = '#,##0.00'

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        file_name = f"yatirimlar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import_excel')
    def import_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "Excel dosyası yüklenmedi."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            data_rows = rows[1:]

            for row in data_rows:
                isin_adi_name = row[0]
                if not isin_adi_name:
                    continue  # Eğer işin adı boşsa bu satırı atla
                isin_adi_instance, _ = IsinAdi.objects.get_or_create(ad=isin_adi_name)

                Yatirim.objects.create(
                    isin_adi=isin_adi_instance,  
                    toplam_kesif_ihale_bedeli=row[1] or 0,
                    yil_sonuna_kadar_harcama=row[2] or 0,
                    yil_kesif_bedeli=row[3] or 0,
                    yil_nakti=row[4] or 0,
                    revize_odenk=row[5] or 0,
                    bbb_ve_sonrasi_kesif_bedeli=row[6] or 0,
                    yapilis_sekli=row[7] or "kamulastırma_disi",
                    baslama_tarihi=row[8] or 2025,
                    bitis_tarihi=row[9] or 2025,
                    talep=row[10] or 0,
                    tenkis=row[11] or 0,
                )

            return Response({"success": "Excel dosyasından yatırımlar başarıyla yüklendi."})
        except Exception as e:
            print(traceback.format_exc())
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class KanunViewSet(viewsets.ModelViewSet):
    queryset = KanunMaddeleri.objects.all().order_by('id')  
    serializer_class = KanunSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = KanunFilter

    def create(self, request, *args, **kwargs):
        dosyalar_data = request.FILES.getlist("dosyalar")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        kanun = serializer.save()

        for dosya in dosyalar_data:
            KanunDosyalari.objects.create(
                kanun_maddesi=kanun,
                dosya=dosya,
                dosya_tipi=dosya.content_type,
            )

        return Response(self.get_serializer(kanun).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        
        dosyalar_data = request.FILES.getlist("dosyalar")
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        for eski_dosya in instance.dosyalar.all():
            if eski_dosya.dosya:
                eski_dosya.dosya.delete(save=False)
            eski_dosya.delete()

        for dosya in dosyalar_data:
            KanunDosyalari.objects.create(
                kanun_maddesi=instance, dosya=dosya, dosya_tipi=dosya.content_type
            )

        return Response(serializer.data)

    @action(detail=True, methods=["delete"], url_path="dosya-sil/(?P<dosya_id>\d+)")
    def dosya_sil(self, request, pk=None, dosya_id=None):
        try:
            dosya = KanunDosyalari.objects.filter(id=dosya_id)
            dosya.delete()
            return Response({"success": "Dosya başarıyla silindi"}, status=status.HTTP_200_OK)
        except KanunDosyalari.DoesNotExist:
            return Response({"detail": "Dosya bulunamadı."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Kanun Maddeleri"

        # Başlık satırı
        headers = [ "Kanun No", "Madde", ]
        ws.append(headers)

        for obj in queryset:
            ws.append([
                obj.no,
                obj.adi,
            ])

        # Tablo ve stil
        last_row = ws.max_row
        last_column = ws.max_column
        table_ref = f"A1:{chr(64 + last_column)}{last_row}"
        tab = Table(displayName="KanunlarTablosu", ref=table_ref)
        ws.add_table(tab)

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_column):
            for cell in row:
                cell.border = border

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"kanunlar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


@api_view(["GET"])
def dashboard(request):
    # 1. Toplam ve çalışan makine sayısı
    makine_sayisi = Makine.objects.count()
    calisan_makine_sayisi = (
        Makine.objects.filter(takipler__isin_durumu="devam").distinct().count()
    )

    # 2. En çok çalışan makine (işe başlama ve bitişi dolu olanları al)
    calisma_sureleri = (
        MakineTakip.objects.filter(ise_baslama__isnull=False, ise_bitis__isnull=False)
        .annotate(
            calisma_suresi=ExpressionWrapper(
                F("ise_bitis") - F("ise_baslama"), output_field=DurationField()
            )
        )
        .values("makina__id", "makina__no", "makina__marka__ad")
        .annotate(toplam_sure=Sum("calisma_suresi"))
        .order_by("-toplam_sure")
        .first()
    )

    if calisma_sureleri:
        en_cok_calisan = {
            "makine_id": calisma_sureleri["makina__id"],
            "makine_adi": calisma_sureleri["makina__no"],
            "makine_marka": calisma_sureleri["makina__marka__ad"],
            "toplam_calisma_gunu": round(
                calisma_sureleri["toplam_sure"].total_seconds() / 86400, 2
            ),
        }
    else:
        en_cok_calisan = None

    # 3. En son başlayan makine takibi
    son_kayit = (
        MakineTakip.objects.filter(ise_baslama__isnull=False)
        .order_by("-ise_baslama")
        .select_related("makina", "is_tanimi")
        .first()
    )

    if son_kayit:
        son_baslayan = {
            "makine_adi": son_kayit.makina.no if son_kayit.makina else None,
            "makine_marka": son_kayit.makina.marka.ad if son_kayit.makina else None,
            "is_tanimi": son_kayit.is_tanimi.tanim if son_kayit.is_tanimi else None,
            "baslama_tarihi": son_kayit.ise_baslama,
            "isin_durumu": son_kayit.isin_durumu,
        }
    else:
        son_baslayan = None

    return Response(
        {
            "toplam_makine_sayisi": makine_sayisi,
            "calisan_makine_sayisi": calisan_makine_sayisi,
            "en_cok_calisan_makine": en_cok_calisan,
            "son_baslayan_makine": son_baslayan,
        }
    )

@api_view(['GET'])
def yil_choices(request):
    current_year = datetime.datetime.now().year
    YIL_CHOICES = [(year, str(year)) for year in range(current_year, current_year + 25)]
    data = [{"value": year, "label": label} for year, label in YIL_CHOICES]
    return Response(data)

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
